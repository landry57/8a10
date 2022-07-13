import os
from unicodedata import name
from colorama import Fore, init, Back, Style
import openpyxl
from flask import Flask,render_template, redirect,url_for,request, send_from_directory,flash
from werkzeug.utils import secure_filename

UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'xls','xlsx'}

app = Flask(__name__)
app.config['SECRET_KEY'] = os.urandom(40)
app.config['MAX_CONTENT_LENGTH'] = 16 * 1000 * 1000
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

prefixOrange=["07", "08", "09", "47", "48", "49", "57", "58", "59", "67", "68", "69", "77", "78", "79", "87", "88", "89", "97" , "98"]
prefixMtn=["04", "05", "06", "44", "45", "46", "54", "55", "56" , "64", "65", "66", "74", "75", "76", "84", "85", "86", "94", "95", "96"]
prefixMoov=["01", "02", "03", "40", "41", "42" , "43", "50", "51", "52", "53", "70", "71", "72", "73"]
prefixMoovFixe=["208", "218", "228", "238"]
prefixOrangeFixe=["202", "203", "212", "213", "215", "217", "224", "225", "234", "235", "243", "244", "245", "306", "316", "319", "327", "337", "347", "359", "368"]
prefixMtnFixe=["200", "210", "220", "230", "240", "300", "310", "320", "330", "340", "350" ,"360"]




@app.route('/')
def index(name=""):
     return render_template('index.html', link=name)


def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route('/send', methods=['GET', 'POST'])
def upload_file():
    if request.method == 'POST':
        rows=request.form['rows']
        cols=request.form['cols']
        if 'file' not in request.files:
            flash('FICHIER REQUIS')
            return redirect(request.url)
        file = request.files['file']
        if rows =="":
             flash('AJOUTER LE NOMBRE DE LINE DU FICHIER')

        if cols =="":
             flash('AJOUTER LA POSITION DE LA COLONNE OU SE TROUVENT LES NUMEROS')             
        if file.filename == '':
            flash('FICHIER REQUIS')
            return redirect(request.url)
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
            wb_obj = openpyxl.load_workbook(os.path.join(app.config['UPLOAD_FOLDER'], filename).strip())
            sheet_obj =  wb_obj.active
            for i in range(2,int(rows)):
                tel= filter(sheet_obj.cell(row=i,column=int(cols)).value)
                sheet_obj.cell(row=i,column=int(cols)).value=tel  
            wb_obj.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
            #sheet_obj.close()
  
           # wb_obj.save()
        
            return redirect(url_for('index', name='/uploads/'+filename))
    return redirect(url_for('index')) 

@app.route('/uploads/<name>')
def download_file(name):
    return send_from_directory(app.config["UPLOAD_FOLDER"], name)


def filter(tel):
    tel=str(tel)
    removePrefix= tel.replace("+225","")
    removePrefix=removePrefix.replace("00225","")
    removePrefix=removePrefix.replace(" ","")
    removePrefix=removePrefix.replace("-","")
    removePrefix=removePrefix.replace("/","")
    if(len(removePrefix)==8):
        pre=removePrefix[0:2]
        if pre in prefixOrange:
            return "+22507"+removePrefix
        if pre in prefixMtn:
            return "+22505"+removePrefix
        if pre in prefixMoov:
            return "+22501"+removePrefix
        if pre in prefixOrangeFixe:
            return "+22527"+removePrefix    
        if pre in prefixMtnFixe:
            return "+22525"+removePrefix       
        if pre in prefixMoovFixe:
            return "+22521"+removePrefix       
  
    return tel
# main driver function
if __name__ == '__main__':
 
    # run() method of Flask class runs the application
    # on the local development server.
    app.run()